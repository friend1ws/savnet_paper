#! /usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side
from string import ascii_uppercase 

wb = Workbook()
wb.guess_types = True

def format_table(work_sheet, input_file, title):
    
    work_sheet["A1"].value = title
    work_sheet["A1"].font = Font(name = "Helvetica", size = 12, bold = True)

    col_len = {} 
    italic_ind = [] 
    with open(input_file, 'r') as hin:
        header = hin.readline().rstrip('\n').split('\t')
        for i in range(len(header)):
            work_sheet[ascii_uppercase[i] + '3'].value = header[i]
            work_sheet[ascii_uppercase[i] + '3'].font = Font(name = "Helvetica", size = 11, bold = True)        
            work_sheet[ascii_uppercase[i] + '3'].border = Border(top = Side(border_style = "medium", color = "FF000000"),
                                                                 bottom = Side(border_style = "medium", color = "FF000000"))

            if header[i] in ["Gene", "Associated splicing factor mutation"]: italic_ind.append(i)

            col = work_sheet[ascii_uppercase[i] + '3'].column
            if col not in col_len: col_len[col] = 0
            col_len[col] = max(col_len[col], len(header[i]))


        row = 0
        for line in hin:
            F = line.rstrip('\n').split('\t')
            for i in range(len(F)):
                work_sheet[ascii_uppercase[i] + str(row + 4)].value = F[i]
                if i in italic_ind:
                    work_sheet[ascii_uppercase[i] + str(row + 4)].font = Font(name = "Helvetica", size = 11, italic = True)
                else:
                    work_sheet[ascii_uppercase[i] + str(row + 4)].font = Font(name = "Helvetica", size = 11)

                col = work_sheet[ascii_uppercase[i] + str(row + 4)].column
                if col not in col_len: col_len[col] = 0
                col_len[col] = max(col_len[col], len(F[i]))

            row = row + 1
        
    def as_text(value): return str(value) if value is not None else ""

    for i in col_len:
        work_sheet.column_dimensions[i].width = float(col_len[i]) * 1.2


ws1 = wb.active
ws1.title = "S1_TCGA_list"
format_table(ws1, "../table/TableS1.txt", "Supplementary Table 1.  Abbreviations for TCGA cancer types")

ws2 = wb.create_sheet("S2_Sample_list")
format_table(ws2, "../table/TableS2.txt", "Supplementary Table 2.  List of samples used in this study and the frequencies of their somatic variants")

ws3 = wb.create_sheet("S3_SAV_list")
format_table(ws3, "../table/TableS3.txt", "Supplementary Table 3.  List of splicing-associated variants (SAV) identified in this study")

ws4 = wb.create_sheet("S4_SF_Mut")
format_table(ws4, "../table/TableS4.txt", "Supplementary Table 4.  List of samples affected by splicing factor mutations")

ws5 = wb.create_sheet("S5_SF_Splice")
format_table(ws5, "../table/TableS5.txt", "Supplementary Table 5.  Significantly associated splicing alterations with splicing factor mutations")

ws6 = wb.create_sheet("S6_Base_Hotspot_list")
format_table(ws6, "../table/TableS6.txt", "Supplementary Table 6.  List of base-level hotspots (shared by ≥ 3 samples) at donor and acceptor sites")

ws7 = wb.create_sheet("S7_SS_Hotspot_list")
format_table(ws7, "../table/TableS7.txt", "Supplementary Table 7.  List of splice site (SS)-level hotspots (shared by ≥ 5 samples) at donor and acceptor sites")

wb.save("../table/Supplementary_Table.xlsx")


